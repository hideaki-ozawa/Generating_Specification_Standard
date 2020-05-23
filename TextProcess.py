import Levenshtein
import re
import openpyxl
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
import os

from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from io import StringIO
from PyPDF2 import PdfFileWriter, PdfFileReader
from tika import parser

from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.tokenize.punkt import PunktSentenceTokenizer, PunktParameters

import tabula


# Preprocess the pdf files.
def convert_pdf_to_txt(path, save_name):
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    laparams = LAParams()
    device = TextConverter(rsrcmgr, retstr, laparams=laparams)
    fp = open(path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0
    caching = True
    pagenos = set()
    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages,
                                  password=password,
                                  caching=caching, check_extractable=True):
        interpreter.process_page(page)
    fp.close()
    device.close()
    str = retstr.getvalue()
    retstr.close()
    try:
        with open("%s" % save_name, "w") as f:
            for i in str:
                f.write(i)
        print("%s Writing Succeed!" % save_name)
    except:
        print("Writing Failed!")


def split_pdf(start_page, end_page, path, save_name):
    output = PdfFileWriter()
    pdf_file = PdfFileReader(open(path, 'rb'))
    for i in range(start_page, end_page):
        output.addPage(pdf_file.getPage(i))
    outputStream = open(save_name, 'wb')
    output.write(outputStream)
    print("%s Writing Succeed!" % save_name)


def split_pdf_endpage(start_page, path, save_name):
    output = PdfFileWriter()
    pdf_file = PdfFileReader(open(path, 'rb'))
    end_page = pdf_file.getNumPages()
    for i in range(start_page, end_page):
        output.addPage(pdf_file.getPage(i))
    outputStream = open(save_name, 'wb')
    output.write(outputStream)
    print("%s Writing Succeed!" % save_name)


def get_pdf_path(file_dir):
    L = []
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            if os.path.splitext(file)[1] == '.pdf':
                L.append(os.path.join(root, file))
    return L


def get_txt_path(file_dir):
    L = []
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            if os.path.splitext(file)[1] == '.txt':
                L.append(os.path.join(root, file))
    return L


def rename_pdf(path):
    filename = get_pdf_path(path)
    for i in range(0, len(filename)):
        newname = filename[i][0:76] + '.pdf'
        os.rename(filename[i], newname)
    L = get_pdf_path(path)
    for i in L:
        print("%s Rename Succeed!" % i)


# Find difference between specifications at the sentence level.
def DifferentWords(Sample, Sentence):
    """
    Return different words in *Sentence* when comparing it to *Sample*.
    :return: ['(+)gas', ... , '(-)diesel', ...]
    """
    stop_words = set(stopwords.words('english'))
    # stop_words = ['i', 'me', 'my', 'myself', 'we', 'our', 'ours', 'ourselves']
    Sample_word_tokens = word_tokenize(Sample)
    Sentence_word_tokens = word_tokenize(Sentence)

    filtered_Sample = [w for w in Sample_word_tokens if not w in stop_words]
    filtered_Sentence = [w for w in Sentence_word_tokens if
                         not w in stop_words]
    new_words = [w for w in filtered_Sentence if not w in filtered_Sample]
    lost_words = [w for w in filtered_Sample if not w in filtered_Sentence]

    for i in range(0, len(new_words)):
        new_words[i] = '(+)' + new_words[i]
    for i in range(0, len(lost_words)):
        lost_words[i] = '(-)' + lost_words[i]

    different_words = new_words + lost_words

    return different_words


def List2Str(list):
    return " & ".join(list)


def Txt2List(path, abbreviation=[]):
    fp = open(path)
    data = fp.read()
    punkt_param = PunktParameters()
    punkt_param.abbrev_types = set(abbreviation)
    tokenizer = PunktSentenceTokenizer(punkt_param)
    text_line = tokenizer.tokenize(data)
    # for sentence in text_line:
    #     if '△' in sentence:
    #         text_line.remove(sentence)
    return text_line


def remove_n(path):
    """
    Remove unexpected \n generated in the transformation from .pdf to .txt.
    """
    f1 = open(path, 'r')
    str = f1.read()
    pattern = re.compile(r'([^\n])\n([^\n])')
    result = pattern.sub(r'\1\2', str)  # 剔除换行符
    f2 = open(path, 'w')
    f2.write(result)
    f2.close()
    f1.close()
    print(path + ' Illegal \\n Removed')


def CreatUnion(union, new_list):
    """
    Create the union set of different sentences.
    :param new_list: A list of sentences splited from a specification.
    :return: ['sentence1', 'sentence2', ...]
    """
    len_u = len(union)
    len_n = len(new_list)
    new_union = union * 1
    mark = 0
    for i in range(0, len_n):
        for j in range(mark, len_u):
            if Levenshtein.distance(new_list[i], union[j]) <= 30:
                mark = j + 1
                break
        else:
            new_union.insert(mark + 1, new_list[i])

    return new_union


def CorrectUnion(union):
    len_u = len(union)
    for i in range(0, len_u):
        for j in range(0, len_u):
            if (Levenshtein.distance(union[i], union[j]) < 10) & (i != j):
                union[j] = '-1'
    return [s for s in union if not s == '-1']


def CompareList(union, new_list):
    """
    Compare the sentences list of specification to the union set. Return '1' when the sentence is in the union set.
    Return different words when the sentence is similar to the sentence in the union set. Return '0' when the list does
    not have the sentence in the union set.
    :param union: A list of sentences containing all types of sentences in specifications.
    :param new_list: A list of sentences splited from a specification.
    :return: ['1', '0', '(+)gas & (-)diesel', ...]
    """
    len_u = len(union)
    len_n = len(new_list)
    result = [None] * (len_u)
    l_distance = [None] * (len_n)
    for i in range(0, len_u):
        for j in range(0, len_n):
            l_distance[j] = Levenshtein.distance(union[i], new_list[j])

        if min(l_distance) == 0:
            result[i] = '1'
        elif min(l_distance) >= 30:
            result[i] = '0'
        else:
            result[i] = List2Str(DifferentWords(union[i], new_list[
                l_distance.index(min(l_distance))]))

    for i in range(0, len_u):
        if (result[i] == '') or (result[i] == '(+) ') or (result[i] == '(-) ') \
                or (result[i] == '(-).') or (result[i] == '(+) '):
            result[i] = '1'

    return result


def subsection_value(shipname,
                     subsection,
                     abbreviation,
                     path,
                     SHIP_TYPE="M9000"
                     ):
    """
    Return the subsection value which will be written into Excel.
    Try to save the following duplicate code:

    list_1727_2_1 = TextProcess.Txt2List('/Users/guichenwei/Documents/test1/txt/General_Part
    /Section2/S1727C9000_2_1.txt',[])
    list_0937_2_1 = TextProcess.Txt2List('/Users/guichenwei/Documents/test1/txt/General_Part
    /Section2/S0937C9000_2_1.txt',[])
    list_Z373_2_1 = TextProcess.Txt2List('/Users/guichenwei/Documents/test1/txt/General_Part
    /Section2/SZ373C9000_2_1.txt',[])

    union_2_1 = TextProcess.CreatUnion(TextProcess.CreatUnion(list_1727_2_1,list_0937_2_1),list_Z373_2_1)
    union_2_1 = TextProcess.CorrectUnion(union_2_1)

    c_1727_2_1 = TextProcess.CompareList(union_2_1, list_1727_2_1)
    c_0937_2_1 = TextProcess.CompareList(union_2_1, list_0937_2_1)
    c_Z373_2_1 = TextProcess.CompareList(union_2_1, list_Z373_2_1)

    union_2_1.insert(0, 'filename')
    c_1727_2_1.insert(0, 'S1727C9000')
    c_0937_2_1.insert(0, 'S0937C9000')
    c_Z373_2_1.insert(0, 'SZ373C9000')

    value_2_1 = [union_2_1, c_1727_2_1, c_0937_2_1, c_Z373_2_1]

    :param shipname: ['1727', '0937', 'Z373', ...]
    :param subsection: '_2_1'
    :param abbreviation: ['reg', 'i.e', 'no', 'incl', ...]
    :param path: '/txt/General_Part/section'
    :return:
    [['filename', 'sentence1', 'sentence2', ...],
     ['S1727C9000', '1', '0', ...              ],
     ...
     ['SK114C9000', '1', '0', ...              ]]
    """

    vn = locals()

    for x in range(len(subsection)):
        book_name_xlsx = './excel/Machinery_Part/Section' + str(x+1) + '.xlsx'
        value = []
        for subsec in subsection[x]:
            for i in range(0, len(shipname)):
                vn['list_' + shipname[i] + subsec] = Txt2List(
                    path + str(x+1) + '/' + subsec + '/S' +
                    shipname[i] + SHIP_TYPE + subsec + '.txt', abbreviation)

            vn['union' + subsec] = vn['list_' + shipname[0] + subsec]
            for i in range(1, len(shipname)):
                vn['union' + subsec] = CreatUnion(vn['union' + subsec], vn['list_' + shipname[i] + subsec])
            vn['union' + subsec] = CorrectUnion(vn['union' + subsec])

            for i in range(0, len(shipname)):
                vn['c_' + shipname[i] + subsec] = CompareList(
                    vn['union' + subsec],
                    vn['list_' + shipname[i] + subsec])

            vn['union' + subsec].insert(0, 'filename')
            for i in range(0, len(shipname)):
                vn['c_' + shipname[i] + subsec].insert(0, 'S' + shipname[i] + SHIP_TYPE)

            vn['value' + subsec] = [vn['union' + subsec]]
            for i in range(0, len(shipname)):
                vn['value' + subsec].append(vn['c_' + shipname[i] + subsec])

            value.append(vn['value' + subsec])

        write_excel_xlsx(book_name_xlsx, subsection[x], value)


# Write the data to Excel.
def write_excel_xlsx(path, sheet_name, value):
    """
    :param path: '.../name.xlsx'
    :param sheet_name: ['1.1', '1.2', ...]
    :param value: Value to be written. Two-dimensional list of the same length as *sheet_name*.
    """
    # sheet.cell(row=i + 1, column=j + 1, value=str(value[0][i][j])) 改变row，column后的i，j，转制
    sheet_number = len(sheet_name)
    workbook = openpyxl.Workbook()
    if sheet_number == 1:
        sheet = workbook.active
        sheet.title = sheet_name[0]
        for i in range(0, len(value[0])):
            for j in range(0, len(value[0][i])):
                # value[0][i][j] = ILLEGAL_CHARACTERS_RE.sub(r'', value[0][i][j])
                sheet.cell(row=j + 1, column=i + 1, value=str(value[0][i][j]))
                # Transposition
                # sheet.cell(row=i + 1, column=j + 1, value=str(value[0][i][j]))
        workbook.save(path)
        print("xlsx格式表格写入数据成功！")
    else:
        sheet = locals()
        sheet0 = workbook.active
        workbook.remove(sheet0)
        for i in range(0, sheet_number):
            sheet[sheet_name[i]] = workbook.create_sheet(sheet_name[i])
        for m in range(0, len(value)):
            for i in range(0, len(value[m])):
                for j in range(0, len(value[m][i])):
                    # value[m][i][j] = ILLEGAL_CHARACTERS_RE.sub(r'', value[m][i][j])
                    sheet[sheet_name[m]].cell(row=j + 1, column=i + 1,
                                              value=str(value[m][i][j]))
        workbook.save(path)
        print(path + '写入数据成功!')


def read_excel_xlsx(path, sheet_name1):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name1]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


def write_matrix_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


# Vectorization
def CompareList_1(union, new_list):
    """
    Compare the sentences list of specification to the union set. Return 1 when the sentence is  similar to
    the sentence in the union set. Return 0 when the list does not have the sentence in the union set.
    :return: [0, 1, 1 ,0, ...]
    """
    len_u = len(union)
    len_n = len(new_list)
    result = [None] * (len_u)
    l_distance = [None] * (len_n)
    for i in range(0, len_u):
        for j in range(0, len_n):
            l_distance[j] = Levenshtein.distance(union[i], new_list[j])

        if min(l_distance) < 30:
            result[i] = 1
        else:
            result[i] = 0

    return result


def cos_distance(matrix):
    """
    Calculate the cosine distance of n arrays.
    :param matrix: [[1, 0, 1,...], [0, 1, 1,...], ...]
    :return: A symmetric matrix.
    """
    dim = len(matrix)
    cols = dim
    rows = dim
    distance = [[0] * cols for i in range(rows)]

    for i in range(0, rows):
        for j in range(0, cols):
            distance[i][j] = np.dot(matrix[i], matrix[j]) \
                             / (np.linalg.norm(matrix[i]) * (
                np.linalg.norm(matrix[j])))
            distance[i][j] = round(distance[i][j], 5)

    return distance


def section_matrix(shipname, subsection, abbreviation, inputpath):
    """
    Return a matrix of 0 and 1.
    Each row represents the result when comparing each specification to the union set.
    :param shipname: ['1727', '0937', 'Z373', ...]
    :param subsection: ['_1_1', '_1_2']
    :param abbreviation: ['reg', 'i.e', 'no', 'incl', ...]
    :param inputpath: '/txt/General_Part/section'
    :return:　[[1, 1, 0, 0,...],       ship 1
               [1, 0, 1, 0,...],       ship 2
               ...
               [1, 1, 1, 1,...]]       ship n
    """
    vn = locals()
    SHIP_TYPE = "M9000"
    section_name = re.search(r'\d+', subsection[0])  # Match object


    for m in range(0, len(subsection)):
        for i in range(0, len(shipname)):
            vn['list_' + shipname[i] + subsection[m]] = Txt2List(
                inputpath + section_name.group() + '/' + subsection[m] + '/S' +
                shipname[i] + SHIP_TYPE + subsection[m] + '.txt', abbreviation)

        vn['union' + subsection[m]] = vn['list_' + shipname[0] + subsection[m]]
        for i in range(1, len(shipname)):
            vn['union' + subsection[m]] = CreatUnion(
                vn['union' + subsection[m]],
                vn['list_' + shipname[i] + subsection[m]])
        vn['union' + subsection[m]] = CorrectUnion(vn['union' + subsection[m]])

        for i in range(0, len(shipname)):
            vn['c_' + shipname[i] + subsection[m]] = CompareList_1(
                vn['union' + subsection[m]],
                vn['list_' + shipname[i] + subsection[
                    m]])

    for i in range(0, len(shipname)):
        vn['value_' + shipname[i] + '_' + subsection[0][1]] = []
    for i in range(0, len(shipname)):
        for m in range(0, len(subsection)):
            value_str = 'value_' + shipname[i] + '_' + subsection[0][1]
            vn[value_str] = vn[value_str] + vn[
                'c_' + shipname[i] + subsection[m]]

    matrix = []
    for i in range(0, len(shipname)):
        matrix.append(vn['value_' + shipname[i] + '_' + subsection[0][1]])

    return matrix


def section_matrix_and_item_union(shipname, subsection, abbreviation,
                                  inputpath):
    """
    Return the outputs of section_matrix() and the sentences union set.
    """
    vn = locals()
    item_union = []
    for m in range(0, len(subsection)):
        for i in range(0, len(shipname)):
            vn['list_' + shipname[i] + subsection[m]] = Txt2List(
                inputpath + subsection[m][1] + '/' + subsection[m] + '/S' +
                shipname[i] + 'M9000' + subsection[m] + '.txt', abbreviation)

        vn['union' + subsection[m]] = vn['list_' + shipname[0] + subsection[m]]
        for i in range(1, len(shipname)):
            vn['union' + subsection[m]] = CreatUnion(
                vn['union' + subsection[m]],
                vn['list_' + shipname[i] + subsection[m]])
        vn['union' + subsection[m]] = CorrectUnion(vn['union' + subsection[m]])

        item_union = item_union + vn['union' + subsection[m]]

        for i in range(0, len(shipname)):
            vn['c_' + shipname[i] + subsection[m]] = CompareList_1(
                vn['union' + subsection[m]],
                vn['list_' + shipname[i] + subsection[m]])

    for i in range(0, len(shipname)):
        vn['value_' + shipname[i] + '_' + subsection[0][1]] = []
    for i in range(0, len(shipname)):
        for m in range(0, len(subsection)):
            value_str = 'value_' + shipname[i] + '_' + subsection[0][1]
            vn[value_str] = vn[value_str] + vn[
                'c_' + shipname[i] + subsection[m]]

    matrix = []
    for i in range(0, len(shipname)):
        matrix.append(vn['value_' + shipname[i] + '_' + subsection[0][1]])

    return matrix, item_union


# def section_matrix_hull(shipname, section, abbreviation, inputpath):
#     vn = locals()
#     for i in range(0, len(shipname)):
#         vn['list_' + shipname[i]] = Txt2List(inputpath + section + '/S' + shipname[i] + 'C9000.txt', abbreviation)
#     union = vn['list_' + shipname[0]]
#     for i in range(1, len(shipname)):
#         union = CreatUnion(union, vn['list_' + shipname[i]])
#     union = CorrectUnion(union)
#     for i in range(0, len(shipname)):
#         vn['value_'+shipname[i]] = CompareList_1(union, vn['list_' + shipname[i]])
#
#     matrix = []
#     for i in range(0, len(shipname)):
#         matrix.append(vn['value_'+shipname[i]])
#     return matrix


def num_difference_array(array1, arrary2):
    array = array1 + arrary2
    num = 0
    for i in range(0, len(array)):
        if array[i] == 1:
            num = num + 1
    return num


def num_difference_matrix(matrix):
    """
    Calculate the number of different items and form the matrix.
    :param matrix: List of outputs of CompareList_1.
    :type numpy.ndarray
    :return: A symmetric matrix which stored number of different items between specifications.
    """
    dim = len(matrix)
    diffrence = [[0 for i in range(dim)] for j in range(dim)]
    for i in range(0, dim):
        for j in range(0, dim):
            diffrence[i][j] = num_difference_array(matrix[i], matrix[j])
    diffrence = np.array(diffrence)
    return diffrence


def squareform(matrix):
    dim = len(matrix)
    squareform_matrix = []
    for i in range(0, dim):
        for j in range(i + 1, dim):
            squareform_matrix.append(matrix[i][j])
    squareform_matrix = np.array(squareform_matrix)
    return squareform_matrix


def split_txt(input_path, sections, save_path):
    '''
    split .txt file(full text) into each section and subsection.

    :param input_path: txt file path
    :param sections(list): m×n dim. m=section, n=subsection
    :param save_path: dir. e.g. "./txt/Machinery_Part"
    :return:
    '''
    print("{0} Split...".format(input_path), end='')

    fp = open(input_path)
    data = fp.read()

    os.makedirs(save_path, exist_ok=True)

    shipname = input_path[-14:-4]
    num_sec = len(sections)
    for i in range(num_sec):
        # split into sections
        sec_path = save_path + "/Section" + str(i + 1)
        start_id = data.find("SECTION {0}".format(i + 1))
        if i != num_sec - 1:
            end_id = data.find("SECTION {0}".format(i + 2))
            sec_data = data[start_id:end_id]
        else:
            sec_data = data[start_id:]

        # split into subsections
        num_subsec = len(sections[i])
        for j in range(num_subsec):
            subsec_ub = "_" + sections[i][j].replace(".", "_")  # e.g _1_1
            subsec_path = sec_path + "/" + subsec_ub
            os.makedirs(subsec_path, exist_ok=True)

            # num of space is different accoding to how to convert pdf to txt
            # start_id = sec_data.find(sections[i][j] + "  ")
            match = re.search(r'[^.]' + sections[i][j] + r'  ', sec_data)
            if match != None:
                start_id = match.start() + 1
            else:
                start_id = -1

            if j != num_subsec - 1:
                # end_id = sec_data.find(sections[i][j + 1] + "  ")
                match = re.search(r'[^.]' + sections[i][j+1] + r'  ', sec_data)
                if match != None:
                    end_id = match.start()
                    subsec_data = sec_data[start_id:end_id]
                else:
                    subsec_data = sec_data[start_id:]
            else:
                subsec_data = sec_data[start_id:]

            subsec_data = subsec_data.replace("\n", ".\n", 1)
            subsec_data = subsec_data.rstrip()
            if subsec_data == "":
                subsec_data = "."
            wf = open(subsec_path + "/" + shipname + subsec_ub + ".txt", "w")
            wf.write(subsec_data)
            wf.close()
    print('Succeed!')


def remove_illegal_char(path):
    """
    Remove unexpected \n generated in the transformation from .pdf to .txt.
    """
    f1 = open(path, 'r')
    data = f1.read()
    pattern = re.compile(
        r'[\000-\010]|[\013-\014]|[\016-\037]')
    # illegal pattern
    # [\000-\010]|[\013-\014]|[\016-\037]|[\x00-\x1f\x7f-\x9f]|[\uffff]
    if isinstance(data, str):
        result = pattern.sub("", data)
    else:
        result = data
    f2 = open(path, 'w')
    f2.write(result)
    f2.close()
    f1.close()
    # print(path + ' Illegal Character Removed')


def add_dot(path):
    """
    Add '.' at the end of senteces
    """
    f1 = open(path, 'r')
    data = f1.read()
    pattern = re.compile('\s*[\n\r](\s*[\n\r])+')
    if isinstance(data, str):
        result = pattern.sub(".\n\n", data)
    else:
        result = data

    pattern = re.compile('[.]{2}\n')
    if isinstance(result, str):
        result = pattern.sub(".\n", result)
    else:
        result = result

    f2 = open(path, 'w')
    f2.write(result)
    f2.close()
    f1.close()
    print(path + ' Specific Character Removed')


def remove_specific_char(path):
    """
    Remove repeated words generated in the transformation from .pdf to .txt.
    e.g "IMABARI SHIPBUILDING CO., LTD", "Ｍ 1 － 1 "
    """
    f1 = open(path, 'r')
    data = f1.read()
    pattern = re.compile(
        r'IMABARI SHIPBUILDING CO., LTD|Ｍ \d* － \d*e?\s*|M\d* － \d*e?\s*')
    if isinstance(data, str):
        result = pattern.sub("", data)
    else:
        result = data
    f2 = open(path, 'w')
    f2.write(result)
    f2.close()
    f1.close()
    print(path + ' Specific Character Removed')


def read_table(pdf_path,
               pages='7-10',
               output_dir='./pdf/table'):
    '''

    :param pdf_path:
    :param pages:str  e.g '7-10', 'all'
    :param output_dir:
    :return:
    '''

    os.makedirs(output_dir, exist_ok=True)

    print("pdf: {0}\npages: {1}".format(pdf_path[-14:], pages))

    output_name = output_dir + '/' + pdf_path[-14:-4] + "_page" + pages + ".txt"
    ln = tabula.read_pdf(pdf_path, lattice=True, pages=pages)
    # page_ln = range(start_page, end_page + 1)
    page_ln = range(len(ln))
    txt = ''
    for i, page in enumerate(page_ln):
        df = ln[i]

        if df.empty:
            pass
        else:
            # todo: Read first line. And replace unname**
            # print(df)
            page_str = column2str(df)
            for index, row in df.iterrows():
                row_str = ' '.join(map(str, row.dropna()))
                row_str = row_str.replace('\r', ' ')
                # print('------------\nrow: {0}'.format(row_str))
                page_str += row_str + '\n\n'

            txt += page_str

    with open(output_name, 'w') as f:
        f.write(txt)


def column2str(df):
    result = ''
    for col in df.columns:
        if 'Unnamed' in col:
            pass
        else:
            result += col.replace('\r', ' ') + ' '
    result += '\n\n'
    # print(result)
    return result


def convert_pdf_to_txt_by_tika(pdf_path, save_name):
    file_data = parser.from_file(pdf_path)
    text = file_data["content"]
    text = text.lstrip()

    try:
        with open("%s" % save_name, "w") as f:
            for i in text:
                f.write(i)
        print("%s Writing Succeed!" % save_name)
    except:
        print("Writing Failed!")


def replace_underbar_period(char):
    '''
    Replace '_1_1' into '1.1'
    :param char: '_1_1'
    :return:
    '''
    char = char[1:]
    result = char.replace('_', '.')
    return result


def search_num(char):
    '''
    :param char: '_1_1' or '1.1'
    :return: [1, 1]
    '''
    result = re.findall('\d+', char)
    return result



def subsection_matrix(shipname, subsection, abbreviation, inputpath):
    '''
    Return a matrix of 0 and 1.
    Each row represents the result when comparing each specification to the union set.
    :param shipname: ['1727', '0937', 'Z373', ...]
    :param subsection: '_1_1'
    :param abbreviation: ['reg', 'i.e', 'no', 'incl', ...]
    :param inputpath: '/txt/General_Part/section'
    :return:　[[1, 1, 0, 0,...],       ship 1
               [1, 0, 1, 0,...],       ship 2
               ...
               [1, 1, 1, 1,...]]       ship n
    '''
    vn = locals()
    SHIP_TYPE = "M9000"
    section_name = re.search(r'\d+', subsection)  # Match object


    for i in range(0, len(shipname)):
        vn['list_' + shipname[i] + subsection] = Txt2List(
            inputpath + section_name.group() + '/' + subsection + '/S' +
            shipname[i] + SHIP_TYPE + subsection + '.txt', abbreviation)

    vn['union' + subsection] = vn['list_' + shipname[0] + subsection]
    for i in range(1, len(shipname)):
        vn['union' + subsection] = CreatUnion(
            vn['union' + subsection],
            vn['list_' + shipname[i] + subsection])
    vn['union' + subsection] = CorrectUnion(vn['union' + subsection])

    for i in range(0, len(shipname)):
        vn['c_' + shipname[i] + subsection] = CompareList_1(
            vn['union' + subsection],
            vn['list_' + shipname[i] + subsection])

    for i in range(0, len(shipname)):
        vn['value_' + shipname[i] + '_' + subsection] = []
    for i in range(0, len(shipname)):
        for m in range(0, len(subsection)):
            value_str = 'value_' + shipname[i] + '_' + subsection
            vn[value_str] = vn[value_str] + vn['c_' + shipname[i] + subsection]

    matrix = []
    for i in range(0, len(shipname)):
        matrix.append(vn['value_' + shipname[i] + '_' + subsection])

    return matrix
